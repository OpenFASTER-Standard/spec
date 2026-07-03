"""Module-agnostic renderer for OpenFASTER modules.

Given a :class:`ModuleConfig` (which points at a module's XSD and its
template-shape callables), this builds the module's:

  * self-documenting Excel workbook (one linked sheet per XSD group),
  * machine-readable metadata store (``*.json``),
  * human-readable field reference (``*.md``), and
  * Bikeshed include (``*.include.bs``) pulled into the specification.

Every data sheet carries four header rows:

    Row 1: technical column name (as in the XSD)
    Row 2: English description (from the XSD ``xs:documentation``)
    Row 3: expected type / format / constraints
    Row 4: Required / Optional / Conditional

Enum columns are rendered as native Excel dropdowns (data validation). All
field-level content is sourced from the XSD by :class:`engine.xsd_model.XsdModel`;
this module only lays it out.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .xsd_model import XsdModel

# Rows 1-4 are the documentation header; data entry starts at row 5.
HEADER_NAME_ROW = 1
HEADER_DESC_ROW = 2
HEADER_TYPE_ROW = 3
HEADER_REQ_ROW = 4
DATA_FIRST_ROW = 5
# Dropdowns / formatting are applied down to this row so future entries inherit
# validation without needing to re-apply it.
DATA_LAST_ROW = 1000

# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #
NAME_FILL = PatternFill("solid", fgColor="1F4E78")
DESC_FILL = PatternFill("solid", fgColor="2E75B6")
TYPE_FILL = PatternFill("solid", fgColor="9DC3E6")
REQ_FILL = PatternFill("solid", fgColor="DEEBF7")

NAME_FONT = Font(bold=True, color="FFFFFF", size=11)
DESC_FONT = Font(italic=True, color="FFFFFF", size=9)
TYPE_FONT = Font(color="1F3864", size=9, bold=True)

REQ_FONTS = {
    "Required": Font(bold=True, color="C00000", size=9),
    "Optional": Font(color="808080", size=9),
    "Conditional": Font(bold=True, color="BF8F00", size=9),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


@dataclass
class ModuleConfig:
    """Everything the engine needs to generate one OpenFASTER module."""

    title: str
    xsd_path: Path
    output_dir: Path
    xlsx_name: str
    json_name: str
    doc_name: str
    bs_name: str
    legend_sheet_name: str
    master_sheet_name: str
    sheet_order: list[str]
    build_enums: Callable[[XsdModel], tuple[dict[str, list[str]], dict[str, dict[str, str]]]]
    build_sheets: Callable[[XsdModel], dict[str, list[tuple]]]
    sheet_info: dict[str, dict[str, str]]
    legend_rows: list[tuple[str, str]]
    link_key: str = "RequestId"


class Generator:
    def __init__(self, config: ModuleConfig):
        self.c = config
        self.model = XsdModel(str(config.xsd_path))
        self.enums, self.enum_meanings = config.build_enums(self.model)
        self.sheets = config.build_sheets(self.model)

    # ------------------------------------------------------------------ #
    # Workbook building
    # ------------------------------------------------------------------ #
    @staticmethod
    def _make_dropdown(values: list[str], lists_ws, list_registry: dict) -> DataValidation:
        """Build a list DataValidation.

        Short value lists are embedded inline; longer ones are stored on the
        hidden ``_Lists`` sheet to stay under Excel's ~255-char inline limit.
        """
        joined = ",".join(values)
        if len(joined) <= 250:
            formula = f'"{joined}"'
        else:
            key = tuple(values)
            if key not in list_registry:
                col = len(list_registry) + 1
                letter = get_column_letter(col)
                lists_ws.cell(row=1, column=col, value=f"list_{col}")
                for i, val in enumerate(values, start=2):
                    lists_ws.cell(row=i, column=col, value=val)
                ref = f"${letter}$2:${letter}${len(values) + 1}"
                list_registry[key] = f"=_Lists!{ref}"
            formula = list_registry[key]

        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,  # False => the dropdown arrow IS shown in Excel
        )
        dv.error = "Please choose a value from the dropdown list."
        dv.errorTitle = "Invalid entry"
        dv.errorStyle = "stop"
        dv.showErrorMessage = True
        return dv

    def _build_sheet(self, ws, fields: list[tuple], lists_ws, list_registry: dict) -> None:
        for col_idx, (name, desc, type_str, req, enum_key) in enumerate(fields, start=1):
            letter = get_column_letter(col_idx)

            name_cell = ws.cell(row=HEADER_NAME_ROW, column=col_idx, value=name)
            name_cell.fill = NAME_FILL
            name_cell.font = NAME_FONT
            name_cell.alignment = WRAP_CENTER
            name_cell.border = BORDER

            desc_cell = ws.cell(row=HEADER_DESC_ROW, column=col_idx, value=desc)
            desc_cell.fill = DESC_FILL
            desc_cell.font = DESC_FONT
            desc_cell.alignment = WRAP_TOP
            desc_cell.border = BORDER

            # For enum columns, show the allowed values on the type row.
            if enum_key:
                type_display = "Choose one: " + " | ".join(self.enums[enum_key])
            else:
                type_display = type_str
            type_cell = ws.cell(row=HEADER_TYPE_ROW, column=col_idx, value=type_display)
            type_cell.fill = TYPE_FILL
            type_cell.font = TYPE_FONT
            type_cell.alignment = WRAP_TOP
            type_cell.border = BORDER

            req_cell = ws.cell(row=HEADER_REQ_ROW, column=col_idx, value=req)
            req_cell.fill = REQ_FILL
            req_cell.font = REQ_FONTS[req]
            req_cell.alignment = WRAP_CENTER
            req_cell.border = BORDER

            # Column width heuristic.
            ws.column_dimensions[letter].width = min(max(len(name) + 3, 18), 42)

            # Attach the dropdown to the whole data-entry range.
            if enum_key:
                dv = self._make_dropdown(self.enums[enum_key], lists_ws, list_registry)
                ws.add_data_validation(dv)
                dv.add(f"{letter}{DATA_FIRST_ROW}:{letter}{DATA_LAST_ROW}")

        ws.row_dimensions[HEADER_NAME_ROW].height = 22
        ws.row_dimensions[HEADER_DESC_ROW].height = 68
        ws.row_dimensions[HEADER_TYPE_ROW].height = 40
        ws.row_dimensions[HEADER_REQ_ROW].height = 18

        # Freeze the four header rows and the link-key column.
        ws.freeze_panes = f"B{DATA_FIRST_ROW}"

    def _build_legend(self, ws) -> None:
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 110

        title = ws.cell(row=1, column=1, value=self.c.title)
        title.font = Font(bold=True, color="1F4E78", size=14)
        ws.cell(row=1, column=2, value=f"Generated from {self.c.xsd_path.name}").font = Font(
            italic=True, color="808080", size=9
        )

        for offset, (label, value) in enumerate(self.c.legend_rows, start=3):
            is_heading = label != "" and value == ""
            label_cell = ws.cell(row=offset, column=1, value=label)
            value_cell = ws.cell(row=offset, column=2, value=value)
            value_cell.alignment = WRAP_TOP
            if is_heading:
                label_cell.font = Font(bold=True, color="FFFFFF", size=11)
                label_cell.fill = PatternFill("solid", fgColor="2E75B6")
            else:
                label_cell.font = Font(bold=True, color="1F3864", size=10)
                label_cell.alignment = Alignment(vertical="top")

        ws.freeze_panes = "A3"

    # ------------------------------------------------------------------ #
    # Metadata store + documentation
    # ------------------------------------------------------------------ #
    def _enum_usage(self) -> dict[str, list[str]]:
        """Map each enum key to the 'Sheet.Field' locations that use it."""
        usage: dict[str, list[str]] = {key: [] for key in self.enums}
        for sheet_name, fields in self.sheets.items():
            for name, _desc, _type_str, _req, enum_key in fields:
                if enum_key:
                    usage[enum_key].append(f"{sheet_name}.{name}")
        return usage

    def _build_metadata(self) -> dict:
        """Collect all field metadata into a serialisable structure."""
        usage = self._enum_usage()

        sheets = []
        for order, (sheet_name, fields) in enumerate(self.sheets.items(), start=1):
            field_list = []
            for position, (name, desc, type_str, req, enum_key) in enumerate(fields, start=1):
                field_list.append({
                    "position": position,
                    "name": name,
                    "description": desc,
                    "type": "Enum" if type_str == "Enum" else type_str,
                    "requiredness": req,
                    "enum": enum_key,
                    "allowedValues": self.enums[enum_key] if enum_key else [],
                })
            info = self.c.sheet_info.get(sheet_name, {})
            sheets.append({
                "order": order,
                "name": sheet_name,
                "significance": info.get("significance", ""),
                "cardinality": info.get("cardinality", ""),
                "whenToFill": info.get("whenToFill", ""),
                "fieldCount": len(field_list),
                "fields": field_list,
            })

        enums = {}
        for key, values in self.enums.items():
            enums[key] = {
                "values": values,
                "meanings": {v: self.enum_meanings.get(key, {}).get(v, "") for v in values},
                "usedIn": usage.get(key, []),
            }

        return {
            "title": self.c.title,
            "generatedFrom": self.c.xsd_path.name,
            "linkKey": self.c.link_key,
            "headerRows": {
                "1": "Technical column name",
                "2": "English description",
                "3": "Expected type / format / constraints",
                "4": "Required / Optional / Conditional",
            },
            "enums": enums,
            "sheets": sheets,
        }

    def _write_metadata_json(self, metadata: dict) -> Path:
        path = self.c.output_dir / self.c.json_name
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(metadata, fh, indent=2, ensure_ascii=False)
        return path

    @staticmethod
    def _anchor(text_value: str) -> str:
        """GitHub-style anchor for a heading."""
        keep = [c for c in text_value.lower() if c.isalnum() or c in " -"]
        return "".join(keep).strip().replace(" ", "-")

    @staticmethod
    def _md_escape(value: str) -> str:
        return value.replace("|", "\\|")

    def _write_documentation_md(self, metadata: dict) -> Path:
        """Render an elaborate Markdown reference from the metadata store."""
        lines: list[str] = []
        add = lines.append
        anchor = self._anchor
        esc = self._md_escape

        add(f"# {metadata['title']} - Documentation")
        add("")
        add(
            f"Auto-generated from [`{self.c.json_name}`]({self.c.json_name}) "
            f"(source schema: `{metadata['generatedFrom']}`). Do not edit by hand - "
            "re-run the generator to refresh."
        )
        add("")

        # ---- Overview ------------------------------------------------------ #
        add("## Overview")
        add("")
        add(
            "This workbook captures MiKaDiv (§45b) third-party disclosures for German "
            "capital income. Each disclosure is spread across several sheets that are all "
            f"tied together by a single key, **{metadata['linkKey']}**. One disclosure = "
            "one RequestId, reused on every sheet that carries data for that disclosure."
        )
        add("")

        add("### How to read each sheet")
        add("")
        add("Every data sheet has four header rows; data entry begins on row 5.")
        add("")
        add("| Row | Meaning |")
        add("| --- | --- |")
        for row_num in sorted(metadata["headerRows"]):
            add(f"| {row_num} | {metadata['headerRows'][row_num]} |")
        add("")

        add("### Requiredness legend")
        add("")
        add("| Value | Meaning |")
        add("| --- | --- |")
        add("| Required | Must always be filled for this record. |")
        add("| Optional | May be left blank. |")
        add("| Conditional | Required only in certain cases - see the field description. |")
        add("")

        add("### Linking model")
        add("")
        add(
            f"- **{metadata['linkKey']}** is the key on `{metadata['sheets'][0]['name']}` "
            "and the first column of every other sheet. It is only used to link the sheets, "
            "so any unique value works (it does not need to be a UUID)."
        )
        add(
            "- **Cancellations:** set `RecordType = Cancel` on the master sheet, fill "
            "`PreviousRequestIdForCancellation` (and optionally `ReportSerialNumber`), and "
            "leave every other sheet empty for that RequestId."
        )
        add(
            "- **Community recipients:** capture a community tax-voucher receiver (up to 10 "
            "members) by setting `ReceiverGroupType = CommunityMember` on the tax voucher "
            "sheets and giving all members of one community the same `CommunityGroupId`."
        )
        add("")

        # ---- Sheets at a glance ------------------------------------------- #
        add("## Sheets at a glance")
        add("")
        add("| Sheet | Fields | Cardinality (rows per RequestId) | When to fill |")
        add("| --- | --- | --- | --- |")
        for sheet in metadata["sheets"]:
            link = f"[{sheet['name']}](#{anchor(sheet['name'])})"
            add(
                f"| {link} | {sheet['fieldCount']} | "
                f"{esc(sheet['cardinality'])} | {esc(sheet['whenToFill'])} |"
            )
        add("")

        # ---- Detailed sheet reference ------------------------------------- #
        add("## Detailed sheet reference")
        add("")
        for sheet in metadata["sheets"]:
            add(f"### {sheet['name']}")
            add("")
            if sheet["significance"]:
                add(f"**Significance.** {sheet['significance']}")
                add("")
            if sheet["cardinality"]:
                add(f"**Cardinality.** {sheet['cardinality']}")
                add("")
            if sheet["whenToFill"]:
                add(f"**When to fill.** {sheet['whenToFill']}")
                add("")
            add("| # | Field | Requiredness | Type / Allowed values | Description |")
            add("| --- | --- | --- | --- | --- |")
            for field_meta in sheet["fields"]:
                if field_meta["allowedValues"]:
                    enum_link = f"[`{field_meta['enum']}`](#{anchor(field_meta['enum'])})"
                    type_col = (
                        f"Enum {enum_link}: "
                        + ", ".join(f"`{v}`" for v in field_meta["allowedValues"])
                    )
                else:
                    type_col = field_meta["type"]
                add(
                    f"| {field_meta['position']} | `{field_meta['name']}` | "
                    f"{field_meta['requiredness']} | "
                    f"{esc(type_col)} | {esc(field_meta['description'])} |"
                )
            add("")

        # ---- Enumerations reference --------------------------------------- #
        add("## Enumerations reference")
        add("")
        add(
            "Every value that can be chosen from a dropdown, with its meaning and the "
            "fields that use it."
        )
        add("")
        for key in metadata["enums"]:
            enum = metadata["enums"][key]
            add(f"### {key}")
            add("")
            if enum["usedIn"]:
                used = ", ".join(f"`{u}`" for u in enum["usedIn"])
                add(f"*Used in:* {used}")
                add("")
            add("| Value | Meaning |")
            add("| --- | --- |")
            for value in enum["values"]:
                meaning = esc(enum["meanings"].get(value, ""))
                add(f"| `{value}` | {meaning} |")
            add("")

        path = self.c.output_dir / self.c.doc_name
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines))
        return path

    # ------------------------------------------------------------------ #
    # Bikeshed include (spec source)
    # ------------------------------------------------------------------ #
    @staticmethod
    def _slug(text_value: str) -> str:
        """Lower-case, hyphenated anchor fragment safe for HTML ids."""
        keep = [c if (c.isalnum() or c == " ") else " " for c in text_value.lower()]
        return "-".join("".join(keep).split())

    @staticmethod
    def _html_escape(value: str) -> str:
        return (
            value.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    def _write_bikeshed_include(self, metadata: dict) -> Path:
        """Render the data dictionary + enumerations as a Bikeshed include.

        The output is consumed verbatim by the specification via
        ``<pre class=include>``. Anchors match the spec cross-references:
        ``#data-dictionary``, ``#enumerations``, ``#sheet-<slug>`` and
        ``#enum-<slug>``.
        """
        lines: list[str] = []
        add = lines.append
        slug = self._slug
        esc = self._html_escape

        add(
            "<!-- Generated by the OpenFASTER engine from "
            f"{self.c.json_name}. Do not edit by hand. -->"
        )
        add("")

        # ---- Data dictionary ---------------------------------------------- #
        add('<h2 id="data-dictionary">Data dictionary</h2>')
        add("")
        add(
            "<p>The following field-level definitions are generated from "
            f"<code>{esc(metadata['generatedFrom'])}</code>. Each group below "
            "corresponds to one sheet of the accompanying Excel template. The "
            f"<code>{esc(metadata['linkKey'])}</code> column links the groups "
            "of a single [=disclosure=] together.</p>"
        )
        add("")

        for sheet in metadata["sheets"]:
            anchor = f"sheet-{slug(sheet['name'])}"
            add(f'<h3 id="{anchor}">{esc(sheet["name"])}</h3>')
            add("")
            if sheet["significance"]:
                add(f"<p><b>Significance.</b> {esc(sheet['significance'])}</p>")
            if sheet["cardinality"]:
                add(f"<p><b>Cardinality.</b> {esc(sheet['cardinality'])}</p>")
            if sheet["whenToFill"]:
                add(f"<p><b>When to fill.</b> {esc(sheet['whenToFill'])}</p>")
            add("")
            add('<table class="complex data longlastcol dictionary">')
            add("  <colgroup>")
            add('    <col style="width:4%">')
            add('    <col style="width:14%">')
            add('    <col style="width:10%">')
            add('    <col style="width:24%">')
            add('    <col style="width:48%">')
            add("  </colgroup>")
            add(
                "  <thead><tr><th>#<th>Field<th>Requiredness"
                "<th>Type / Allowed values<th>Description</tr></thead>"
            )
            add("  <tbody>")
            for field_meta in sheet["fields"]:
                if field_meta["allowedValues"]:
                    enum_anchor = f"enum-{slug(field_meta['enum'])}"
                    values = ", ".join(
                        f"<code>{esc(v)}</code>" for v in field_meta["allowedValues"]
                    )
                    type_col = (
                        f'Enum (<a href="#{enum_anchor}">{esc(field_meta["enum"])}</a>): '
                        f"{values}"
                    )
                else:
                    type_col = esc(field_meta["type"])
                add(
                    f"    <tr><td>{field_meta['position']}"
                    f"<td><code>{esc(field_meta['name'])}</code>"
                    f"<td>{esc(field_meta['requiredness'])}"
                    f"<td>{type_col}"
                    f'<td class="long">{esc(field_meta["description"])}</tr>'
                )
            add("  </tbody>")
            add("</table>")
            add("")

        # ---- Enumerations ------------------------------------------------- #
        add('<h2 id="enumerations">Enumerations</h2>')
        add("")
        add(
            "<p>Every value that an enum-typed field may carry, with its meaning and "
            "the fields that use it.</p>"
        )
        add("")
        for key, enum in metadata["enums"].items():
            anchor = f"enum-{slug(key)}"
            add(f'<h3 id="{anchor}">{esc(key)}</h3>')
            add("")
            if enum["usedIn"]:
                used = ", ".join(f"<code>{esc(u)}</code>" for u in enum["usedIn"])
                add(f"<p><em>Used in:</em> {used}</p>")
                add("")
            add('<table class="complex data longlastcol enum-table">')
            add("  <colgroup>")
            add('    <col style="width:18%">')
            add('    <col style="width:82%">')
            add("  </colgroup>")
            add("  <thead><tr><th>Value<th>Meaning</tr></thead>")
            add("  <tbody>")
            for value in enum["values"]:
                meaning = esc(enum["meanings"].get(value, ""))
                add(f'    <tr><td><code>{esc(value)}</code><td class="long">{meaning}</tr>')
            add("  </tbody>")
            add("</table>")
            add("")

        path = self.c.output_dir / self.c.bs_name
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines))
        return path

    # ------------------------------------------------------------------ #
    # Orchestration
    # ------------------------------------------------------------------ #
    def run(self) -> None:
        self.c.output_dir.mkdir(parents=True, exist_ok=True)

        # Write the documentation outputs first so they refresh even if the
        # workbook file is currently open in Excel (Windows locks open files).
        metadata = self._build_metadata()
        print(f"Wrote {self._write_metadata_json(metadata)} (field metadata store)")
        print(f"Wrote {self._write_documentation_md(metadata)} (generated documentation)")
        print(f"Wrote {self._write_bikeshed_include(metadata)} (Bikeshed include)")

        wb = Workbook()
        wb.remove(wb.active)  # remove the default sheet; we create all explicitly

        legend_ws = wb.create_sheet(self.c.legend_sheet_name)
        self._build_legend(legend_ws)

        # Hidden sheet that backs any long dropdown lists.
        lists_ws = wb.create_sheet("_Lists")
        list_registry: dict = {}

        for sheet_name in self.c.sheet_order:
            ws = wb.create_sheet(sheet_name)
            self._build_sheet(ws, self.sheets[sheet_name], lists_ws, list_registry)

        lists_ws.sheet_state = "hidden"
        wb.move_sheet(lists_ws, offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_Lists"))
        wb.active = wb.sheetnames.index(self.c.master_sheet_name)

        xlsx_path = self.c.output_dir / self.c.xlsx_name
        try:
            wb.save(xlsx_path)
            print(f"Wrote {xlsx_path} with sheets: {', '.join(wb.sheetnames)}")
        except PermissionError:
            print(
                f"WARNING: could not write {xlsx_path} - it looks like the file is open. "
                "Close it in Excel and re-run to regenerate the workbook. "
                "(The JSON and documentation were still updated.)"
            )
