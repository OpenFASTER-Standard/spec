from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent


def test_generate_template_produces_seven_sheets_plus_legend_and_meta():
    subprocess.run([sys.executable, "generate_template.py"], cwd=ROOT, check=True)
    wb = openpyxl.load_workbook(ROOT / "kafe" / "generated" / "kafe-v4.3.2.xlsx")
    names = wb.sheetnames
    assert names[0] == "0 Legend Notes"
    assert names[-2] == "_Lists"  # hidden, second-to-last per engine/generator.py:601-603
    assert names[-1] == "Meta"
    # The 6 real data sheets (Meta/Legend/_Lists are not among production's 7 --
    # production's 7th sheet, "meta", is what engine/generator.py's own Meta
    # sheet replaces/represents here) are present:
    for expected in ["1 Creditors Natural", "2 Creditors Juridical",
                      "3 Certificates Of Residence", "4 Income",
                      "5 Investment Chain", "6 Transaction Data"]:
        assert expected in names


def test_income_sheet_has_no_record_type_column():
    wb = openpyxl.load_workbook(ROOT / "kafe" / "generated" / "kafe-v4.3.2.xlsx")
    ws = wb["4 Income"]
    header_names = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
    assert "RecordType" not in header_names
